#coding = utf-8


import openpyxl

wb1 = openpyxl.load_workbook(r'C:\Users\wutao\Desktop\work\户籍.xlsx')
wb2 = openpyxl.load_workbook(r'C:\Users\wutao\Desktop\work\志愿者名.xlsx')
ws1 = wb1.worksheets[0]
ws2 = wb2.worksheets[0]


rows_data = list(ws1.rows)
columns_data = list(ws2.columns)

titles = []
for title in rows_data[0]:
    titles.append(title.value)
print(titles)
tuple = []
for case in rows_data[1:]:
    tem = []
    for cell in case:
        tem.append(cell.value)
    tuple.append(tem)

#print(tuple)
#print(rows_data)

#print(dict)

column = columns_data[0]

result = []
#print(nametup)
for col in column:
    tup = []
    for data in tuple:
        #print(data)
        if col.value in data:
            tup.append(data)
    l = len(tup)
    if l == 1:
        result.append(tup[0])
    elif l > 1:
        for t in tup:
            print(t)
        s = input("请输入：")
        result.append(tup[int(s)])
    else:
        print(col.value+"不存在!")
    '''
    if col.value in dict:
        print(col.value)
        result.append(dict[col.value][3])
    else:
        result.append("户籍册中不存在")
    '''

#print(result)
#print(result[0][0])
for i in range(len(result)):
    #print(type(result[i][0]))
    ws2.cell(i+1,1).value = result[i][0]
    if len(result[i][5]) > 0:
        ws2.cell(i+1,2).value = result[i][5]

wb2.save(r'C:\Users\wutao\Desktop\work\target3.xlsx')
wb1.close()
wb2.close()

